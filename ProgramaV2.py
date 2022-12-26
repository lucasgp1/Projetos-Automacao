from openpyxl import Workbook, load_workbook
import pyodbc
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import datetime as dt

# Realizar conexão com banco de dados SQL
dados_conexao = (
    "Driver={SQL SERVER};"
    "Server=LucasMachine;"
    "Database=ALGAR_RELATORIO_GERAL;"
)

def abre_planilha(detentora_site, nome_site, codigo_site, cidade_site, estado_site, endereco_site, latitude_site, longitude_site):

    detentora_site = str(detentora_site)

    if detentora_site == "['SBA']":

        tabela = load_workbook(r"C:\Users\Lucas\OneDrive\Área de Trabalho\Projetos Algar\Modelos FSCIs\FSCI-Modelo-SBA.xlsx")
        aba_ativa = tabela.active
        data_criacao = dt.datetime.now()
        data_criacao = data_criacao.strftime("%d/%m/%Y")

        aba_ativa["D12"] = str(nome_site)
        aba_ativa["D5"] = str(data_criacao)
        aba_ativa["H12"] = str(codigo_site)[2:-2]
        aba_ativa["L12"] = str(cidade_site)[2:-2]
        aba_ativa["H14"] = str(estado_site)[2:-2]
        aba_ativa["K9"] = str(endereco_site)[2:-2]
        aba_ativa["D14"] = str(latitude_site)[2:-2]
        aba_ativa["D15"] = str(longitude_site)[2:-2]

        tabela.save(rf"C:\Users\Lucas\OneDrive\Área de Trabalho\Projetos Algar\Modelos FSCIs\Planilha_FSCI_{nome_site}.xlsx")
        messagebox.showwarning("PLANILHA CRIADA", "SALVO COM SUCESSO")

    if detentora_site == "['AMERICAN TOWER']":

        tabela = load_workbook(r"C:\Users\Lucas\OneDrive\Área de Trabalho\Projetos Algar\Modelos FSCIs\FSCI-Modelo-ATC.xlsx")
        aba_ativa = tabela.active
        data_criacao = dt.datetime.now()
        data_criacao = data_criacao.strftime("%d/%m/%Y")

        aba_ativa["D12"] = str(nome_site)
        aba_ativa["D5"] = str(data_criacao)
        aba_ativa["H12"] = str(codigo_site)[2:-2]
        aba_ativa["L12"] = str(cidade_site)[2:-2]
        aba_ativa["H14"] = str(estado_site)[2:-2]
        aba_ativa["K9"] = str(endereco_site)[2:-2]
        aba_ativa["D14"] = str(latitude_site)[2:-2]
        aba_ativa["D15"] = str(longitude_site)[2:-2]

        tabela.save(rf"C:\Users\Lucas\OneDrive\Área de Trabalho\Projetos Algar\Modelos FSCIs\Planilha_FSCI_{nome_site}.xlsx")
        messagebox.showwarning("PLANILHA CRIADA", "SALVO COM SUCESSO")

    else:
        messagebox.showwarning("NÃO TEMOS ESSE MODELO DE FSCI", "NÃO TEMOS ESSE MODELO DE FSCI")

def puxa_banco():

    conexao = pyodbc.connect(dados_conexao)
    print("Conexão bem sucedida")
    cursor = conexao.cursor()

    try:
        nome_site = entry_descricao.get()

        consulta = f"""SELECT * FROM  [dbo].['DADOS DETENTORAS$']
        WHERE ESTACAO_SEM_CODIGO = '{nome_site}'"""

        cursor.execute(consulta)

        linhas = cursor.fetchall()

        print(linhas)

        for linha in linhas:
            detentora_site = [linha[1]]
            codigo_site = [linha[2]]
            cidade_site = [linha[3]]
            estado_site = [linha[4]]
            endereco_site = [linha[5]]
            latitude_site = [linha[6]]
            longitude_site = [linha[7]]

        for linha in linhas:
            print("A DETENTORA DO SITE {} é: {}".format(nome_site,linha[1]))

        abre_planilha(detentora_site, nome_site, codigo_site, cidade_site, estado_site, endereco_site, latitude_site,longitude_site)

    except:
        messagebox.showerror("Site Não Encontrado","Site Não Encontrado")

def center(win):

    win.update_idletasks()
    width = win.winfo_width()
    frm_width = win.winfo_rootx() - win.winfo_x()
    win_width = width + 2 * frm_width
    height = win.winfo_height()
    titlebar_height = win.winfo_rooty() - win.winfo_y()
    win_height = height + titlebar_height + frm_width
    x = win.winfo_screenwidth() // 2 - win_width // 2
    y = win.winfo_screenheight() // 2 - win_height // 2
    win.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    win.deiconify()


janela = tk.Tk()

janela.attributes('-alpha', 0.0) # Opcional, para deixar a janela totalmente transparente até os ajustes serem feitos.
janela.minsize(300, 80)
center(janela) # A função
janela.attributes('-alpha', 1.0) # A interface fica visível novamente.


janela.title('Consulta Detentora')

label_nomesite = tk.Label(text="Nome do site:")
label_nomesite.grid(row=1, column=1,padx=100, pady=10, sticky='nswe', columnspan =4)

entry_descricao = tk.Entry()
entry_descricao.grid(row=2,column=1, padx=80, pady=10, sticky='nswe', columnspan =4)

botao_realizar_consulta = tk.Button(text="Realizar Consulta", command=puxa_banco)
botao_realizar_consulta.grid(row=5,column=1,padx = 70, pady=10,sticky='nswe', columnspan =4)

janela.mainloop()
