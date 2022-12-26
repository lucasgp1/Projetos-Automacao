from openpyxl import Workbook, load_workbook
import pyodbc
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import datetime as dt

# Realizar conexão com banco de dados SQL
dados_conexao = (
    "Driver={SQL SERVER};"
    "Server=LucasMachine;"
    "Database=ALGAR_RELATORIO_GERAL;"
)

conexao = pyodbc.connect(dados_conexao)
print("Conexão bem sucedida")
cursor = conexao.cursor()

comando_puxa_antenas = """SELECT [Modelos]
FROM [ALGAR_RELATORIO_GERAL].[dbo].['MODELOS DE EQUIPAMENTOS$']
WHERE [Modelos] IS NOT NULL"""



cursor.execute(comando_puxa_antenas)
lista_modelos_antenas = cursor.fetchall()


print(len(lista_modelos_antenas))
print(lista_modelos_antenas)
lista_modelos_rrus = ["FXED", "FRPA", "RRU5502"]
lista_nova = []


for antena in lista_modelos_antenas:
    antena = [antena[0]]
    lista_nova.append(antena)

print(lista_nova)

def gera_planilha(detentora_site, nome_site, codigo_site, cidade_site, estado_site, endereco_site, latitude_site, longitude_site):

    detentora_site = str(detentora_site)

    comando_puxa_dadosdomodelo = f"""SELECT * FROM [dbo].['MODELOS DE EQUIPAMENTOS$'] WHERE [Modelos] = '{combobox_tipomodeloequipamento1.get()}'"""
    cursor.execute(comando_puxa_dadosdomodelo)
    linhas_antena_selecionada = cursor.fetchall()
    print(linhas_antena_selecionada)

    for caracteristicas_antenas in linhas_antena_selecionada:
        fabricante = [caracteristicas_antenas[0]]
        print(fabricante)
        dimensoes = [caracteristicas_antenas[6]]


    print(type(fabricante))
    print(dimensoes)

    if detentora_site == "['SBA']":

        tabela = load_workbook(r"C:\Users\Lucas\OneDrive\Área de Trabalho\Projetos Algar\Modelos FSCIs\FSCI-Modelo-SBA.xlsx")
        aba_ativa = tabela.active
        data_criacao = dt.datetime.now()
        data_criacao = data_criacao.strftime("%d/%m/%Y")

        aba_ativa["D12"] = str(nome_site)
        aba_ativa["D5"] = str(data_criacao)
        aba_ativa["H12"] = str(codigo_site)[2:-2]
        aba_ativa["L12"] = str(cidade_site)[2:-2]
        aba_ativa["H14"] = str(estado_site)[2:-2]
        aba_ativa["K9"] = str(endereco_site)[2:-2]
        aba_ativa["D14"] = str(latitude_site)[2:-2]
        aba_ativa["D15"] = str(longitude_site)[2:-2]

        tabela.save(rf"C:\Users\Lucas\OneDrive\Área de Trabalho\Projetos Algar\Modelos FSCIs\Planilha_FSCI_{nome_site}.xlsx")
        messagebox.showwarning("PLANILHA CRIADA", "SITE {} SALVO COM SUCESSO")

    elif detentora_site == "['AMERICAN TOWER']":

        tabela = load_workbook(r"C:\Users\Lucas\OneDrive\Área de Trabalho\Projetos Algar\Modelos FSCIs\FSCI-Modelo-ATC.xlsx")
        aba_ativa = tabela.active
        data_criacao = dt.datetime.now()
        data_criacao = data_criacao.strftime("%d/%m/%Y")

        aba_ativa["D12"] = str(nome_site)
        aba_ativa["D5"] = str(data_criacao)
        aba_ativa["H12"] = str(codigo_site)[2:-2]
        aba_ativa["L12"] = str(cidade_site)[2:-2]
        aba_ativa["H14"] = str(estado_site)[2:-2]
        aba_ativa["K9"] = str(endereco_site)[2:-2]
        aba_ativa["D14"] = str(latitude_site)[2:-2]
        aba_ativa["D15"] = str(longitude_site)[2:-2]
        aba_ativa["B25"] = combobox_tiposolicitacao1.get()
        aba_ativa["C25"] = combobox_tipoequipamento1.get()
        aba_ativa["E25"] = combobox_tipomodeloequipamento1.get()
        aba_ativa["D25"] = str(fabricante)[2:-2]
        aba_ativa["G25"] = str(dimensoes)[2:-2]
        aba_ativa["F25"] = str("Necessário Implementar Freq. Operação")

        tabela.save(rf"C:\Users\Lucas\OneDrive\Área de Trabalho\Projetos Algar\Modelos FSCIs\Planilha_FSCI_{nome_site}.xlsx")
        messagebox.showwarning("PLANILHA CRIADA", f"SITE {nome_site} SALVO COM SUCESSO")
        entry_descricao.delete(0, "end")
        entry_descricao.insert(0, "")

    else:
        messagebox.showwarning("NÃO TEMOS ESSE MODELO DE FSCI", "NÃO TEMOS ESSE MODELO DE FSCI")

def puxa_banco():

    try:
        print("Entra aqui")
        nome_site = entry_descricao.get()

        consulta = f"""SELECT * FROM  [dbo].['DADOS DETENTORAS$']
            WHERE ESTACAO_SEM_CODIGO = '{nome_site}'"""

        cursor.execute(consulta)
        linhas = cursor.fetchall()

        print(linhas)

        for linha in linhas:
            detentora_site = [linha[1]]
            codigo_site = [linha[2]]
            cidade_site = [linha[3]]
            estado_site = [linha[4]]
            endereco_site = [linha[5]]
            latitude_site = [linha[6]]
            longitude_site = [linha[7]]

        for linha in linhas:
            print("A DETENTORA DO SITE {} é: {}".format(nome_site, linha[1]))

        gera_planilha(detentora_site, nome_site, codigo_site, cidade_site, estado_site, endereco_site, latitude_site,
                      longitude_site)

    except:
        messagebox.showerror("Site Não Encontrado", "Site Não Encontrado")

def center(win):

    win.update_idletasks()
    width = win.winfo_width()
    frm_width = win.winfo_rootx() - win.winfo_x()
    win_width = width + 2 * frm_width
    height = win.winfo_height()
    titlebar_height = win.winfo_rooty() - win.winfo_y()
    win_height = height + titlebar_height + frm_width
    x = win.winfo_screenwidth() // 2 - win_width // 2
    y = win.winfo_screenheight() // 2 - win_height // 2
    win.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    win.deiconify()

def pega_antenas_rrus(event):

    if selected_equipamento1.get() == "RRU":
        combobox_tipomodeloequipamento1.config(value=lista_modelos_rrus)
    else:
        combobox_tipomodeloequipamento1.config(value=lista_nova)

    combobox_tipomodeloequipamento1.current(0)

    if selected_equipamento2.get() == "RRU":
        combobox_tipomodeloequipamento2.config(value=lista_modelos_rrus)
    else:
        combobox_tipomodeloequipamento2.config(value=lista_nova)

    combobox_tipomodeloequipamento2.current(0)

    if selected_equipamento3.get() == "RRU":
        combobox_tipomodeloequipamento3.config(value=lista_modelos_rrus)
    else:
        combobox_tipomodeloequipamento3.config(value=lista_nova)

    combobox_tipomodeloequipamento3.current(0)

    if selected_equipamento4.get() == "RRU":
        combobox_tipomodeloequipamento4.config(value=lista_modelos_rrus)
    else:
        combobox_tipomodeloequipamento4.config(value=lista_nova)

    combobox_tipomodeloequipamento4.current(0)

    if selected_equipamento5.get() == "RRU":
        combobox_tipomodeloequipamento5.config(value=lista_modelos_rrus)
    else:
        combobox_tipomodeloequipamento5.config(value=lista_nova)

    combobox_tipomodeloequipamento5.current(0)



lista_tipos_solicitacao = ["INSTALAÇÃO","REMOÇÃO"]
lista_tipos_equipamento = ["RRU", "RF"]


janela = tk.Tk()

janela.attributes('-alpha', 0.0) # Opcional, para deixar a janela totalmente transparente até os ajustes serem feitos.
janela.minsize(1200, 500)
center(janela) # A função
janela.attributes('-alpha', 1.0) # A interface fica visível novamente.


janela.title('Programa para Gerar FSCIs')

label_nomesite = tk.Label(text="Nome do site:")
label_nomesite.grid(row=1, column=0,padx=20, pady=10, sticky='nswe', columnspan =4)

entry_descricao = tk.Entry()
entry_descricao.grid(row=1,column=4, padx=20, pady=10, sticky='nswe', columnspan =4)

botao_realizar_consulta = tk.Button(text="Realizar Consulta", command=puxa_banco)
botao_realizar_consulta.grid(row=3,column=1,padx = 20, pady=10,sticky='nswe', columnspan =4)

label_tipo_solicitacao1 = tk.Label(text="Tipo de Solicitação(1):")
label_tipo_solicitacao2 = tk.Label(text="Tipo de Solicitação(2):")
label_tipo_solicitacao3 = tk.Label(text="Tipo de Solicitação(3):")
label_tipo_solicitacao4 = tk.Label(text="Tipo de Solicitação(4):")
label_tipo_solicitacao5 = tk.Label(text="Tipo de Solicitação(5):")
label_tipo_solicitacao1.grid(row=4, column=0,padx=20, pady=10, sticky='nswe', columnspan =4)
label_tipo_solicitacao2.grid(row=5, column=0,padx=20, pady=10, sticky='nswe', columnspan =4)
label_tipo_solicitacao3.grid(row=6, column=0,padx=20, pady=10, sticky='nswe', columnspan =4)
label_tipo_solicitacao4.grid(row=7, column=0,padx=20, pady=10, sticky='nswe', columnspan =4)
label_tipo_solicitacao5.grid(row=8, column=0,padx=20, pady=10, sticky='nswe', columnspan =4)

combobox_tiposolicitacao1 = ttk.Combobox(values=lista_tipos_solicitacao)
combobox_tiposolicitacao2 = ttk.Combobox(values=lista_tipos_solicitacao)
combobox_tiposolicitacao3 = ttk.Combobox(values=lista_tipos_solicitacao)
combobox_tiposolicitacao4 = ttk.Combobox(values=lista_tipos_solicitacao)
combobox_tiposolicitacao5 = ttk.Combobox(values=lista_tipos_solicitacao)
combobox_tiposolicitacao1.grid(row=4, column=4, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tiposolicitacao2.grid(row=5, column=4, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tiposolicitacao3.grid(row=6, column=4, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tiposolicitacao4.grid(row=7, column=4, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tiposolicitacao5.grid(row=8, column=4, padx = 20, pady=10, sticky='nswe', columnspan = 2)

label_tipo_equipamento1 = tk.Label(text="Tipo de Equipamento:")
label_tipo_equipamento2 = tk.Label(text="Tipo de Equipamento:")
label_tipo_equipamento3 = tk.Label(text="Tipo de Equipamento:")
label_tipo_equipamento4 = tk.Label(text="Tipo de Equipamento:")
label_tipo_equipamento5 = tk.Label(text="Tipo de Equipamento:")
label_tipo_equipamento1.grid(row=4, column=6,padx=20, pady=10, sticky='nswe', columnspan =4)
label_tipo_equipamento2.grid(row=5, column=6,padx=20, pady=10, sticky='nswe', columnspan =4)
label_tipo_equipamento3.grid(row=6, column=6,padx=20, pady=10, sticky='nswe', columnspan =4)
label_tipo_equipamento4.grid(row=7, column=6,padx=20, pady=10, sticky='nswe', columnspan =4)
label_tipo_equipamento5.grid(row=8, column=6,padx=20, pady=10, sticky='nswe', columnspan =4)

selected_equipamento1 = tk.StringVar()
selected_equipamento2 = tk.StringVar()
selected_equipamento3 = tk.StringVar()
selected_equipamento4 = tk.StringVar()
selected_equipamento5 = tk.StringVar()
combobox_tipoequipamento1 = ttk.Combobox(values=lista_tipos_equipamento, textvariable=selected_equipamento1)
combobox_tipoequipamento2 = ttk.Combobox(values=lista_tipos_equipamento, textvariable=selected_equipamento2)
combobox_tipoequipamento3 = ttk.Combobox(values=lista_tipos_equipamento, textvariable=selected_equipamento3)
combobox_tipoequipamento4 = ttk.Combobox(values=lista_tipos_equipamento, textvariable=selected_equipamento4)
combobox_tipoequipamento5 = ttk.Combobox(values=lista_tipos_equipamento, textvariable=selected_equipamento5)
combobox_tipoequipamento1.grid(row=4, column=10, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipoequipamento2.grid(row=5, column=10, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipoequipamento3.grid(row=6, column=10, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipoequipamento4.grid(row=7, column=10, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipoequipamento5.grid(row=8, column=10, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipoequipamento1.bind('<<ComboboxSelected>>', pega_antenas_rrus)
combobox_tipoequipamento2.bind('<<ComboboxSelected>>', pega_antenas_rrus)
combobox_tipoequipamento3.bind('<<ComboboxSelected>>', pega_antenas_rrus)
combobox_tipoequipamento4.bind('<<ComboboxSelected>>', pega_antenas_rrus)
combobox_tipoequipamento5.bind('<<ComboboxSelected>>', pega_antenas_rrus)

label_modelo_equipamento1 = tk.Label(text="Modelo do Equipamento:")
label_modelo_equipamento2 = tk.Label(text="Modelo do Equipamento:")
label_modelo_equipamento3 = tk.Label(text="Modelo do Equipamento:")
label_modelo_equipamento4 = tk.Label(text="Modelo do Equipamento:")
label_modelo_equipamento5 = tk.Label(text="Modelo do Equipamento:")
label_modelo_equipamento1.grid(row=4, column=14,padx=20, pady=10, sticky='nswe', columnspan =4)
label_modelo_equipamento2.grid(row=5, column=14,padx=20, pady=10, sticky='nswe', columnspan =4)
label_modelo_equipamento3.grid(row=6, column=14,padx=20, pady=10, sticky='nswe', columnspan =4)
label_modelo_equipamento4.grid(row=7, column=14,padx=20, pady=10, sticky='nswe', columnspan =4)
label_modelo_equipamento5.grid(row=8, column=14,padx=20, pady=10, sticky='nswe', columnspan =4)

combobox_tipomodeloequipamento1 = ttk.Combobox(values=[" "])
combobox_tipomodeloequipamento2 = ttk.Combobox(values=[" "])
combobox_tipomodeloequipamento3 = ttk.Combobox(values=[" "])
combobox_tipomodeloequipamento4 = ttk.Combobox(values=[" "])
combobox_tipomodeloequipamento5 = ttk.Combobox(values=[" "])
combobox_tipomodeloequipamento1.grid(row=4, column=18, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipomodeloequipamento2.grid(row=5, column=18, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipomodeloequipamento3.grid(row=6, column=18, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipomodeloequipamento4.grid(row=7, column=18, padx = 20, pady=10, sticky='nswe', columnspan = 2)
combobox_tipomodeloequipamento5.grid(row=8, column=18, padx = 20, pady=10, sticky='nswe', columnspan = 2)

janela.mainloop()

